import streamlit as st
import pandas as pd
import numpy as np
import numpy_financial as npf
import openpyxl
import requests
import io
from io import BytesIO
from datetime import datetime

st.set_page_config(page_title="DCF Project Calculator", layout="wide", page_icon="📊")

# ===== MÁRGENES / PADDING DE LA APP (edita estos valores) =====
PAGE_MAX_WIDTH    = "90vw"   # ancho máximo del contenido (acepta vw, px, %, etc.)
PAGE_PADDING_TOP  = "1.5rem" # espacio arriba, antes del título
PAGE_PADDING_LEFT  = "1rem"  # margen izquierdo del contenido
PAGE_PADDING_RIGHT = "1rem"  # margen derecho del contenido

# ===== ANCHO DE ETIQUETAS DE TÍTULO =====
# Las tablas ocupan el 100% del ancho disponible (use_container_width), así que
# los títulos usan el mismo 100% para quedar alineados con ellas.
SECTION_HDR_WIDTH  = "100%"  # ancho de los títulos de sección (INFLOWS, OUTFLOWS, TOTALES, etc.)
SUBGROUP_HDR_WIDTH = "100%"  # ancho de las etiquetas de subgrupo (REVENUE, COSTS & EXPENSES, etc.)

st.markdown(f"""
<style>
section[data-testid="stSidebar"] {{ display: none; }}
.main .block-container {{
    padding-top: {PAGE_PADDING_TOP};
    max-width: {PAGE_MAX_WIDTH} !important;
    margin-left: auto !important;
    margin-right: auto !important;
    padding-left: {PAGE_PADDING_LEFT};
    padding-right: {PAGE_PADDING_RIGHT};
}}
.kpi-card {{
    background: #ffffff; border-radius: 10px; padding: 14px 10px;
    box-shadow: 0 2px 8px rgba(0,0,0,0.10); text-align: center;
    margin-bottom: 12px; min-height: 110px;
    display: flex; flex-direction: column; justify-content: center; align-items: center;
}}
.kpi-label {{ font-size: 10px; font-weight: 700; color: #666; text-transform: uppercase; letter-spacing: 0.5px; line-height: 1.4; }}
.kpi-val   {{ font-size: 20px; font-weight: 900; color: #374151; margin: 6px 0 3px; }}
.kpi-sub   {{ font-size: 10px; color: #999; }}
.kpi-val-green {{ font-size: 20px; font-weight: 900; color: #1F2937; margin: 6px 0 3px; }}
.section-hdr {{
    font-size: 13px; font-weight: 800; color: #333333;
    letter-spacing: 2px; text-transform: uppercase;
    background: #FEFBDF; border-left: 4px solid #FEFBDF; padding: 6px 10px; margin: 20px 0 6px;
    width: {SECTION_HDR_WIDTH}; box-sizing: border-box;
}}
.subgroup-hdr {{
    background: #FEFBDF; padding: 5px 12px 5px 16px;
    font-size: 11px; font-weight: 800; letter-spacing: 1.2px;
    color: #333333; border-left: 4px solid #FEFBDF;
    margin: 6px 0 1px 0;
    width: {SUBGROUP_HDR_WIDTH}; box-sizing: border-box;
}}
.page-title {{ font-size: 30px; font-weight: 900; color: #111827; letter-spacing: 1px; text-align: center; margin-bottom: 1.2em; }}
.page-sub   {{ font-size: 13px; color: #888; }}
</style>
""", unsafe_allow_html=True)

# ===== ANCHOS DE TABLAS (edita estos valores) =====
TABLES_USE_FIXED_WIDTH = False   # True = usar anchos fijos definidos abajo, False = ocupar todo el ancho disponible
CONCEPT_COL_WIDTH = 220         # ancho columna "Concepto" en px (tablas INFLOWS/OUTFLOWS/FINANCING/TOTALES/FCF)
YEAR_COL_WIDTH    = 100         # ancho de cada columna de año y SUBTOTAL en px (mismas tablas)

st.markdown("""
<style>
div[data-testid="stDataFrame"],
div[data-testid="stDataFrame"] > div,
.stDataFrameGlideDataEditor {
    width: 100% !important;
}
</style>
""", unsafe_allow_html=True)

@st.cache_data(ttl=300)
def _download_xlsx():
    FILE_ID = st.secrets["FILE_ID"]
    r = requests.get(f"https://docs.google.com/spreadsheets/d/{FILE_ID}/export?format=xlsx")
    return r.content

@st.cache_data
def get_projects():
    wb = openpyxl.load_workbook(io.BytesIO(_download_xlsx()), read_only=True)
    return [s for s in wb.sheetnames if s not in ("INSTRUCCIONES", "PLANTILLA")]

@st.cache_data
def load_defaults(project_name: str):
    wb = openpyxl.load_workbook(io.BytesIO(_download_xlsx()), data_only=True)
    ws = wb[project_name]

    header = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
    years = []
    for v in header[2:]:
        try:
            y = int(v)
            if 1900 < y < 2200:
                years.append(y)
        except (TypeError, ValueError):
            pass
    years = years[:10]  # limitar a un horizonte de 10 años
    n = len(years)

    KNOWN = {"INFLOWS", "OUTFLOWS", "FINANCING"}
    sections = {"INFLOWS": [], "OUTFLOWS": [], "FINANCING": []}
    current = None

    def _f(v):
        try:
            return float(v)
        except (TypeError, ValueError):
            return 0.0

    for row in ws.iter_rows(min_row=3, max_row=39, values_only=True):
        sec     = str(row[0]).strip() if row[0] else ""
        concept = str(row[1]).strip() if row[1] else ""

        if sec in KNOWN:
            current = sec
            if concept:
                vals = [_f(v) for v in row[2:2 + n]]
                sections[current].append((concept, vals))
            continue

        if current and concept:
            vals = [_f(v) for v in row[2:2 + n]]
            sections[current].append((concept, vals))

    DEFAULTS = {
        "INFLOWS":   ["Revenue 1", "Revenue 2"],
        "OUTFLOWS":  ["CAPEX", "OPEX", "Comm 1", "Comm 2", "Otro Gasto 1", "Otro Gasto 2"],
        "FINANCING": ["Debt Draw", "Principal Repayment", "Interest"],
    }

    for sec, default_concepts in DEFAULTS.items():
        loaded = {r[0]: r[1] for r in sections[sec]}
        ordered = []
        for c in default_concepts:
            ordered.append((c, loaded.get(c, [0.0] * n)))
        for c, v in sections[sec]:
            if c not in default_concepts:
                ordered.append((c, v))
        sections[sec] = ordered

    return sections, years

def fmt_usd(v):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return f"(${abs(v):,.0f})" if v < 0 else f"${v:,.0f}"

def fmt_pct(pct, none_str="—"):
    # pct ya viene en escala de porcentaje (ej. 10.0 = 10%), no en fracción.
    if pct is None or (isinstance(pct, float) and np.isnan(pct)):
        return none_str
    return f"({abs(pct):.2f}%)" if pct < 0 else f"{pct:.2f}%"

def _pdf_safe(s):
    # Helvetica (fuente core del PDF) solo soporta Latin-1; cualquier
    # carácter fuera de ese rango (tipografía "inteligente" pegada desde
    # Word/Excel, emojis, etc.) hace que fpdf lance una excepción y la
    # descarga falle en silencio. Lo sustituimos en vez de reventar.
    return str(s).encode("latin-1", "replace").decode("latin-1")

# Sub-group visual grouping (purely display — data structure unchanged)
CONCEPT_SUBGROUP = {
    "Revenue 1":  "REVENUE",
    "Revenue 2":  "REVENUE",
    "CAPEX":      "COSTS & EXPENSES",
    "OPEX":       "COSTS & EXPENSES",
    "Comm 1":     "COMMISSIONS",
    "Comm 2":     "COMMISSIONS",
}
SEC_SUBGROUPS = {
    "INFLOWS":   ["REVENUE"],
    "OUTFLOWS":  ["COSTS & EXPENSES", "OTROS", "COMMISSIONS"],
    "FINANCING": ["FCF FROM FINANCING"],
}
SEC_DEFAULT_SG = {
    "INFLOWS":   "REVENUE",
    "OUTFLOWS":  "OTROS",
    "FINANCING": "FCF FROM FINANCING",
}

def kpi_card(label, value, sub="", green=False):
    # fmt_usd()/fmt_pct() envuelven los negativos entre paréntesis: rojo.
    # Los valores positivos van en azul. "—" (sin dato) se deja neutro.
    val_str = value.strip() if isinstance(value, str) else str(value)
    is_negative = val_str.startswith("(") and val_str.endswith(")")
    is_empty    = val_str in ("—", "-")
    if is_negative:
        color = "#DC2626"
    elif is_empty:
        color = None
    else:
        color = "#1D4ED8"
    cls = "kpi-val-green" if green else "kpi-val"
    style = f' style="color:{color}"' if color else ""
    return (f'<div class="kpi-card"><div class="kpi-label">{label}</div>'
            f'<div class="{cls}"{style}>{value}</div><div class="kpi-sub">{sub}</div></div>')

# Ancho de columna para "Concepto" y para columnas numéricas (años / SUBTOTAL)
_CONCEPT_W = CONCEPT_COL_WIDTH if TABLES_USE_FIXED_WIDTH else "medium"
_YEAR_W    = YEAR_COL_WIDTH if TABLES_USE_FIXED_WIDTH else "small"
TABLES_USE_CONTAINER_WIDTH = not TABLES_USE_FIXED_WIDTH

def col_cfg(scols):
    cfg = {"Concepto": st.column_config.TextColumn("Concepto", width=_CONCEPT_W, pinned=True)}
    cfg.update({
        y: st.column_config.NumberColumn(f"Año {i + 1}", format="$%,.0f", width=_YEAR_W)
        for i, y in enumerate(scols)
    })
    cfg["SUBTOTAL"] = st.column_config.NumberColumn("SUBTOTAL", format="$%,.0f", width=_YEAR_W)
    return cfg

# Colores de formato condicional: valores negativos en rojo, positivos en
# negro; la etiqueta (columna Concepto) en azul, o en rojo si el total de
# esa fila es negativo.
CLR_POS   = "#000000"
CLR_NEG   = "#DC2626"
CLR_LABEL = "#1D4ED8"

def _row_style_fn(r):
    subtotal = r.get("SUBTOTAL", 0)
    lbl_color = CLR_NEG if subtotal < 0 else CLR_LABEL
    styles = []
    for col in r.index:
        if col == "Concepto":
            styles.append(f"background-color:#FEFBDF;color:{lbl_color};font-weight:bold")
        else:
            val = r[col]
            color = CLR_NEG if (isinstance(val, (int, float)) and val < 0) else CLR_POS
            styles.append(f"background-color:#FEFBDF;color:{color};font-weight:bold")
    return styles

def total_row_style(df, num_cols):
    return df.style.apply(_row_style_fn, axis=1).format(
        lambda x: "-" if x == 0 else (f"({abs(x):,.0f})" if x < 0 else f"${x:,.0f}"), subset=num_cols
    )

def sum_by_year(section, n):
    return [sum(vals[i] for _, vals in section) for i in range(n)]

def concept_total(section, name):
    for concept, vals in section:
        if concept == name:
            return sum(vals)
    return 0

def render_fcf_row(label, year_vals, subtotal, scols):
    row = {"Concepto": label}
    row.update({scols[i]: year_vals[i] for i in range(len(scols))})
    row["SUBTOTAL"] = subtotal
    df = pd.DataFrame([row])
    st.dataframe(
        df.style.apply(_row_style_fn, axis=1).format(
            lambda x: f"({abs(x):,.0f})" if x < 0 else f"${x:,.0f}",
            subset=scols + ["SUBTOTAL"],
        ),
        use_container_width=TABLES_USE_CONTAINER_WIDTH,
        hide_index=True,
        column_config=col_cfg(scols),
    )

def render_section(title, key, section_data, scols, selected):
    st.markdown(f'<div class="section-hdr">{title}</div>', unsafe_allow_html=True)

    # Assign each concept to its sub-group
    sg_data = {}
    for concept, vals in section_data:
        sg = CONCEPT_SUBGROUP.get(concept, SEC_DEFAULT_SG.get(key, "OTHER"))
        sg_data.setdefault(sg, []).append((concept, vals))

    subgroups  = SEC_SUBGROUPS.get(key, ["OTHER"])
    all_results = []

    for sg in subgroups:
        sg_concepts = sg_data.get(sg, [])
        if not sg_concepts:
            continue

        st.markdown(f'<div class="subgroup-hdr">{sg}</div>', unsafe_allow_html=True)

        n_rows   = len(sg_concepts)
        vals_key = f"vals_{key}_{sg}_{selected}"

        if vals_key not in st.session_state or len(st.session_state[vals_key]) != n_rows:
            st.session_state[vals_key] = [list(c[1]) for c in sg_concepts]

        labels     = [c[0] for c in sg_concepts]
        row_totals = [sum(v) for v in st.session_state[vals_key]]

        df = pd.DataFrame(st.session_state[vals_key], columns=scols)
        df.insert(0, "Concepto", labels)
        df["SUBTOTAL"] = row_totals

        edited = st.data_editor(
            df,
            use_container_width=TABLES_USE_CONTAINER_WIDTH,
            num_rows="fixed",
            key=f"editor_{key}_{sg}_{selected}",
            disabled=["SUBTOTAL"],
            column_config=col_cfg(scols),
            hide_index=True,
        )

        edited[scols] = edited[scols].fillna(0).astype(float)

        sg_results    = []
        new_vals_list = []
        for i in range(len(edited)):
            concept = str(edited.iloc[i]["Concepto"] or sg_concepts[i][0])
            vals    = edited.iloc[i][scols].tolist()
            sg_results.append((concept, vals))
            new_vals_list.append(vals)

        if new_vals_list != st.session_state[vals_key]:
            st.session_state[vals_key] = new_vals_list
            st.rerun()

        # Consolidated total row for this sub-group
        sg_year_totals = edited[scols].sum()
        sg_total       = sg_year_totals.sum()
        total_row = pd.DataFrame([{
            "Concepto": f"▶ {sg}",
            **sg_year_totals.to_dict(),
            "SUBTOTAL": sg_total,
        }])
        st.dataframe(
            total_row_style(total_row, list(scols) + ["SUBTOTAL"]),
            use_container_width=TABLES_USE_CONTAINER_WIDTH,
            hide_index=True,
            column_config=col_cfg(scols),
        )

        all_results.extend(sg_results)

    return all_results

st.markdown('<div class="page-title">Modelo dinámico de valoración y análisis financiero</div>', unsafe_allow_html=True)
st.markdown('<div class="page-sub">Selecciona un proyecto · edita las celdas · los resultados se recalculan automáticamente</div>', unsafe_allow_html=True)

projects = get_projects()
col_sel, col_rate = st.columns([3, 2])
with col_sel:
    selected = st.selectbox("Proyecto", projects, key="project_selector")
with col_rate:
    discount_rate_pct = st.number_input(
        "Tasa de Descuento (%) WACC", min_value=0.0, max_value=100.0,
        value=10.0, step=0.5, format="%.2f", key="discount_rate",
    )
discount_rate = discount_rate_pct / 100.0

D, YEARS = load_defaults(selected)
SCOLS = [str(y) for y in YEARS]
N = len(YEARS)
st.divider()

equity_container  = st.container()
metrics_container = st.container()
st.divider()

inflows  = render_section("INFLOWS",  "INFLOWS",  D["INFLOWS"],  SCOLS, selected)
outflows = render_section("OUTFLOWS", "OUTFLOWS", D["OUTFLOWS"], SCOLS, selected)

capex_amount = abs(concept_total(outflows, "CAPEX"))

# Depreciación lineal del CAPEX, según la vida útil del activo. Es un gasto
# contable (no representa una salida de caja): no se suma a OUTFLOWS ni
# afecta FCF/TIR/VAN. Solo se usa para calcular EBITDA/EBIT más abajo.
st.markdown('<div class="section-hdr">DEPRECIACIÓN</div>', unsafe_allow_html=True)
useful_life_years = st.number_input(
    "Vida Útil del Activo (años)", min_value=1, max_value=50,
    value=10, step=1, key="useful_life_years",
)
dep_years          = min(int(useful_life_years), N)
annual_dep         = capex_amount / useful_life_years if useful_life_years else 0.0
depreciation_yr    = [annual_dep if i < dep_years else 0.0 for i in range(N)]
depreciation_total = sum(depreciation_yr)
render_fcf_row("Depreciación Lineal (CAPEX / Vida Útil)", depreciation_yr, depreciation_total, SCOLS)

with equity_container:
    st.markdown('<div class="section-hdr">ESTRUCTURA DE CAPITAL</div>', unsafe_allow_html=True)

    col_preopex, col_pct = st.columns([1, 1])
    with col_preopex:
        preopex_amount = st.number_input(
            "PreOpex (Capital Operativo Pre-apertura)", min_value=0.0,
            value=0.0, step=1000.0, format="%.0f", key="preopex_amount",
        )
    with col_pct:
        financing_pct = st.number_input(
            "% Financiamiento", min_value=0.0, max_value=100.0,
            value=70.0, step=5.0, format="%.1f", key="financing_pct",
        )

    total_project_cost = capex_amount + preopex_amount
    equity_pct          = 100.0 - financing_pct
    equity_amount        = total_project_cost * equity_pct / 100.0
    financing_amount     = total_project_cost * financing_pct / 100.0

    col_total, col_equity, col_fin, col_preopex_kpi = st.columns([1, 1, 1, 1])
    with col_total:
        st.markdown(kpi_card("Costo Total del Proyecto", fmt_usd(total_project_cost), "CAPEX + PreOpex"), unsafe_allow_html=True)
    with col_equity:
        st.markdown(kpi_card("EQUITY", fmt_usd(equity_amount), f"{equity_pct:.1f}% del Costo Total"), unsafe_allow_html=True)
    with col_fin:
        st.markdown(kpi_card("FINANCIAMIENTO", fmt_usd(financing_amount), f"{financing_pct:.1f}% del Costo Total", green=True), unsafe_allow_html=True)
    with col_preopex_kpi:
        st.markdown(kpi_card("PreOpex (Capital Operativo Pre-apertura)", fmt_usd(preopex_amount)), unsafe_allow_html=True)

inflows_yr  = sum_by_year(inflows, N)
outflows_yr = sum_by_year(outflows, N)
fcf_no_fin  = [inflows_yr[i] + outflows_yr[i] for i in range(N)]
npv_no      = sum(fcf_no_fin)

# Totales consolidados INFLOWS / OUTFLOWS
st.markdown('<div class="section-hdr">TOTALES</div>', unsafe_allow_html=True)
totals_df = pd.DataFrame([
    {"Concepto": "▶ TOTAL INFLOWS",  **{SCOLS[i]: inflows_yr[i]  for i in range(N)}, "SUBTOTAL": sum(inflows_yr)},
    {"Concepto": "▶ TOTAL OUTFLOWS", **{SCOLS[i]: outflows_yr[i] for i in range(N)}, "SUBTOTAL": sum(outflows_yr)},
])
st.dataframe(
    totals_df.style.apply(_row_style_fn, axis=1).format(
        lambda x: "-" if x == 0 else (f"({abs(x):,.0f})" if x < 0 else f"${x:,.0f}"),
        subset=SCOLS + ["SUBTOTAL"],
    ),
    use_container_width=TABLES_USE_CONTAINER_WIDTH,
    hide_index=True,
    column_config=col_cfg(SCOLS),
)

render_fcf_row("FCF (Excluye Financiamiento)", fcf_no_fin, npv_no, SCOLS)

financing    = render_section("FINANCING", "FINANCING", D["FINANCING"], SCOLS, selected)
financing_yr = sum_by_year(financing, N)
fcf_with_fin = [fcf_no_fin[i] + financing_yr[i] for i in range(N)]
npv_fin      = sum(fcf_with_fin)

render_fcf_row("FCF (Incluye Financiamiento)", fcf_with_fin, npv_fin, SCOLS)

# ===== AMORTIZACIÓN DE PRÉSTAMO (Método Francés) =====
# Calculadora independiente de referencia: no se conecta al FCF/TIR/VAN ni a
# la sección FINANCING. Cuota constante, capital creciente e interés
# decreciente período a período.
st.divider()
st.markdown('<div class="section-hdr">AMORTIZACIÓN DE PRÉSTAMO (Método Francés)</div>', unsafe_allow_html=True)

col_loan_amt, col_loan_rate, col_loan_periods = st.columns(3)
with col_loan_amt:
    loan_amount = st.number_input(
        "Monto del Préstamo", min_value=0.0, value=0.0, step=10000.0,
        format="%.0f", key="loan_amount",
    )
with col_loan_rate:
    loan_rate_pct = st.number_input(
        "Tasa de Interés Anual (%)", min_value=0.0, max_value=100.0,
        value=10.0, step=0.5, format="%.2f", key="loan_rate_pct",
    )
with col_loan_periods:
    loan_periods = st.number_input(
        "Número de Amortizaciones (años)", min_value=1, max_value=N,
        value=N, step=1, key="loan_periods",
    )

loan_rate    = loan_rate_pct / 100.0
loan_periods = int(loan_periods)

def french_amortization(principal, rate, periods):
    if periods <= 0 or principal <= 0:
        return [0.0] * periods, [0.0] * periods, [0.0] * periods
    cuota = principal / periods if rate == 0 else principal * rate / (1 - (1 + rate) ** -periods)
    balance = principal
    capital_l, interest_l, cuota_l = [], [], []
    for _ in range(periods):
        interest = balance * rate
        capital  = cuota - interest
        balance -= capital
        capital_l.append(capital)
        interest_l.append(interest)
        cuota_l.append(capital + interest)
    return capital_l, interest_l, cuota_l

capital_yr, interest_yr, cuota_yr = french_amortization(loan_amount, loan_rate, loan_periods)
loan_scols = [f"L{i + 1}" for i in range(loan_periods)]

amort_df = pd.DataFrame([
    {"Concepto": "Capital",           **{loan_scols[i]: capital_yr[i]  for i in range(loan_periods)}, "SUBTOTAL": sum(capital_yr)},
    {"Concepto": "Interés",           **{loan_scols[i]: interest_yr[i] for i in range(loan_periods)}, "SUBTOTAL": sum(interest_yr)},
    {"Concepto": "Capital + Interés", **{loan_scols[i]: cuota_yr[i]    for i in range(loan_periods)}, "SUBTOTAL": sum(cuota_yr)},
])
st.dataframe(
    amort_df.style.apply(_row_style_fn, axis=1).format(
        lambda x: "-" if x == 0 else (f"({abs(x):,.0f})" if x < 0 else f"${x:,.0f}"),
        subset=loan_scols + ["SUBTOTAL"],
    ),
    use_container_width=TABLES_USE_CONTAINER_WIDTH,
    hide_index=True,
    column_config=col_cfg(loan_scols),
)

def safe_irr(cf):
    try:
        v = npf.irr(cf)
        if v is None:
            return None
        v = float(np.real(v))
        return v if not np.isnan(v) else None
    except Exception:
        return None

irr_no  = safe_irr(fcf_no_fin)
irr_fin = safe_irr(fcf_with_fin)

def safe_npv(rate, cf):
    try:
        return float(npf.npv(rate, cf))
    except Exception:
        return None

van_no  = safe_npv(discount_rate, fcf_no_fin)
van_fin = safe_npv(discount_rate, fcf_with_fin)

def safe_cagr(vals):
    # Tasa de crecimiento anual compuesta, calculada sobre los mismos
    # valores anuales que la fila TOTAL INFLOWS de la tabla de totales.
    # Usa el primer y el último año con Revenue > 0 (ignora años en $0 al
    # inicio -por construcción- o al final -tras cerrar/vender el proyecto-).
    nonzero = [i for i, v in enumerate(vals) if v > 0]
    if len(nonzero) < 2:
        return None
    start_i, end_i = nonzero[0], nonzero[-1]
    years_span = end_i - start_i
    if years_span <= 0:
        return None
    try:
        return (vals[end_i] / vals[start_i]) ** (1 / years_span) - 1
    except Exception:
        return None

noi_last   = inflows_yr[-1] + outflows_yr[-1]
sales_last = inflows_yr[-1]
cap_rate   = noi_last / sales_last if sales_last != 0 else 0

equity_actual = sum(-fcf_with_fin[i] for i in range(N) if fcf_with_fin[i] < 0)
cash_on_cash  = npv_fin / equity_actual if equity_actual != 0 else None

with metrics_container:
    st.markdown('<div class="section-hdr">INVESTMENT RETURNS — Métricas Clave</div>', unsafe_allow_html=True)

    sales_total        = sum(inflows_yr)
    opex_total         = abs(concept_total(outflows, "OPEX"))
    net_cash_generated = npv_fin
    roi                = (net_cash_generated / equity_amount * 100) if equity_amount else None
    revenue_cagr       = safe_cagr(inflows_yr)

    ebitda_total = sales_total - opex_total
    ebit_total   = ebitda_total - depreciation_total

    roi_label  = fmt_pct(roi)
    cagr_label = fmt_pct(revenue_cagr * 100 if revenue_cagr is not None else None)

    irr_no_label  = fmt_pct(irr_no  * 100 if irr_no  is not None else None)
    irr_fin_label = fmt_pct(irr_fin * 100 if irr_fin is not None else None)

    # Grid CSS explícito (en vez de filas independientes de st.columns) para
    # garantizar que cada tarjeta quede exactamente debajo de la de arriba,
    # sin desalinearse por tarjetas de distinta altura (p. ej. subtítulos largos).
    col_table, col_kpis = st.columns([2, 2])
    with col_table:
        st.markdown(f"""
        <div style="background:#ECFDF5;border-radius:12px;padding:16px;height:520px;box-sizing:border-box;">
        <div style="display:grid;grid-template-columns:1fr 1fr;column-gap:12px;">
            {kpi_card("SALES", fmt_usd(sales_total))}
            {kpi_card("CAGR Revenue", cagr_label, "Crecimiento anual compuesto (todo el período)")}
            {kpi_card("CAPEX", fmt_usd(capex_amount))}
            {kpi_card("OPEX", fmt_usd(opex_total))}
            {kpi_card("EBIT", fmt_usd(ebit_total), "EBITDA - Depreciación")}
            {kpi_card("EBITDA", fmt_usd(ebitda_total), "Revenue - OPEX")}
            {kpi_card("NET CASH GENERATED", fmt_usd(net_cash_generated))}
        </div>
        </div>
        """, unsafe_allow_html=True)
    with col_kpis:
        st.markdown(f"""
        <div style="background:#EFF6FF;border-radius:12px;padding:16px;height:520px;box-sizing:border-box;">
        <div style="display:grid;grid-template-columns:1fr 1fr;column-gap:12px;">
            {kpi_card("TIR Sin Financiamiento", irr_no_label)}
            {kpi_card("VAN Sin Financiamiento", fmt_usd(van_no), f"{discount_rate_pct:.1f}%")}
            {kpi_card("TIR Con Financiamiento", irr_fin_label)}
            {kpi_card("VAN Con Financiamiento", fmt_usd(van_fin), f"{discount_rate_pct:.1f}%", green=True)}
            {kpi_card("ROI", roi_label)}
        </div>
        </div>
        """, unsafe_allow_html=True)

st.divider()
st.markdown('<div class="section-hdr">DESCARGAR REPORTE</div>', unsafe_allow_html=True)
fmt_choice = st.radio("Selecciona el formato:", ["Excel (.xlsx)", "PDF (.pdf)"], horizontal=True)

def build_excel():
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
        wb_out   = writer.book

        BG_HEADER = "#FEFBDF"   # Encabezados y subtítulos de sección (único fondo amarillo)
        BG_SUB    = "#FFFFFF"   # Filas de subtotal/total -> blanco
        BG_REG    = "#FFFFFF"   # Filas regulares -> blanco

        hdr_fmt  = wb_out.add_format({"bold": True, "bg_color": BG_HEADER, "font_color": "#333333", "border": 1, "align": "center"})
        title_fmt = wb_out.add_format({"bold": True, "font_size": 14, "font_color": "#333333"})
        sub_fmt  = wb_out.add_format({"bold": True, "bg_color": BG_HEADER, "font_color": "#333333", "border": 1, "indent": 1})
        date_fmt = wb_out.add_format({"italic": True, "font_size": 10, "font_color": "#888888"})

        # Formato condicional: valores numéricos negativos -> rojo, positivos -> negro.
        # Las etiquetas (texto, no numérico) siempre en negro.
        _fmt_cache = {}

        def get_num_fmt(bg, color, bold=False):
            key = ("num", bg, color, bold)
            if key not in _fmt_cache:
                _fmt_cache[key] = wb_out.add_format({
                    "num_format": '#,##0;(#,##0)', "border": 1, "align": "right",
                    "bg_color": bg, "font_color": color, "bold": bold,
                })
            return _fmt_cache[key]

        def get_lbl_fmt(bg, color, bold=False):
            key = ("lbl", bg, color, bold)
            if key not in _fmt_cache:
                _fmt_cache[key] = wb_out.add_format({
                    "bold": bold, "bg_color": bg, "border": 1, "indent": 1, "font_color": color,
                })
            return _fmt_cache[key]

        def get_pct_fmt(bg, color, bold=True):
            key = ("pct", bg, color, bold)
            if key not in _fmt_cache:
                _fmt_cache[key] = wb_out.add_format({
                    "num_format": '0.00%;(0.00%)', "border": 1, "bg_color": bg,
                    "align": "right", "font_color": color, "bold": bold,
                })
            return _fmt_cache[key]

        ws = wb_out.add_worksheet("DCF PROJECT")
        writer.sheets["DCF PROJECT"] = ws
        ws.set_column(0, 0, 32)
        for c in range(1, N + 3):
            ws.set_column(c, c, 14)

        ncols = N + 1
        ts = datetime.now().strftime("%d/%m/%Y  %H:%M")
        ws.merge_range(0, 0, 0, ncols, f"Generado: {ts}", date_fmt)
        ws.merge_range(1, 0, 1, ncols, f"Modelo dinámico de valoración y análisis financiero — {selected.upper()}", title_fmt)
        year_labels = [f"Año {i + 1}" for i in range(len(SCOLS))]
        ws.write(2, 0, "Concepto", hdr_fmt)
        for c, h in enumerate(year_labels + ["SUBTOTAL"], 1):
            ws.write(2, c, h, hdr_fmt)

        rn = 3

        def write_sec(title, rows, total=None):
            nonlocal rn
            ws.merge_range(rn, 0, rn, ncols, title, sub_fmt)
            rn += 1
            for label, vals in rows:
                row_sum = sum(vals)
                ws.write(rn, 0, label, get_lbl_fmt(BG_REG, CLR_POS, bold=True))
                for c, v in enumerate(vals, 1):
                    color = CLR_NEG if v < 0 else CLR_POS
                    ws.write(rn, c, v, get_num_fmt(BG_REG, color))
                sub_color = CLR_NEG if row_sum < 0 else CLR_POS
                ws.write(rn, len(vals) + 1, row_sum, get_num_fmt(BG_SUB, sub_color, bold=True))
                rn += 1
            if total is not None:
                t_label, t_vals = total
                write_total_row(t_label, t_vals)
            rn += 1

        def write_total_row(label, vals):
            nonlocal rn
            row_sum = sum(vals)
            ws.write(rn, 0, label, get_lbl_fmt(BG_SUB, CLR_POS, bold=True))
            for c, v in enumerate(vals, 1):
                color = CLR_NEG if v < 0 else CLR_POS
                ws.write(rn, c, v, get_num_fmt(BG_SUB, color, bold=True))
            sub_color = CLR_NEG if row_sum < 0 else CLR_POS
            ws.write(rn, len(vals) + 1, row_sum, get_num_fmt(BG_SUB, sub_color, bold=True))
            rn += 1

        write_sec("INFLOWS",  inflows,  total=("TOTAL INFLOWS",  inflows_yr))
        write_sec("OUTFLOWS", outflows, total=("TOTAL OUTFLOWS", outflows_yr))

        # Depreciación: gasto contable (no es salida de caja), no se suma a
        # OUTFLOWS ni al FCF; se muestra aparte y solo alimenta el EBIT.
        write_total_row("DEPRECIACIÓN (Lineal)", depreciation_yr)
        rn += 1

        # Línea independiente FCF Project (Inflow - Outflow), separada de las
        # secciones vecinas por filas en blanco.
        write_total_row("FCF PROJECT", fcf_no_fin)
        rn += 1

        write_sec("FCF FROM FINANCING", financing, total=("TOTAL FINANCING", financing_yr))

        ws.merge_range(rn, 0, rn, ncols, "FREE CASH FLOW", sub_fmt); rn += 1
        write_total_row("FCF (Sin Financiamiento)", fcf_no_fin)
        write_total_row("FCF (Con Financiamiento)", fcf_with_fin)
        rn += 1

        ws.merge_range(rn, 0, rn, ncols, "INVESTMENT RETURNS", sub_fmt); rn += 1
        ws.write(rn, 0, "Tasa de Descuento", get_lbl_fmt(BG_SUB, CLR_POS, bold=True))
        ws.write(rn, 1, discount_rate, get_pct_fmt(BG_SUB, CLR_POS))
        rn += 1
        for label, val, is_pct in [
            ("IRR Sin Financiamiento", irr_no  or 0, True),
            ("IRR Con Financiamiento", irr_fin or 0, True),
            ("NPV Sin Financiamiento", van_no  or 0, False),
            ("NPV Con Financiamiento", van_fin or 0, False),
            ("EBITDA",                 ebitda_total,       False),
            ("Depreciación",           depreciation_total, False),
            ("EBIT",                   ebit_total,         False),
        ]:
            val_color = CLR_NEG if val < 0 else CLR_POS
            ws.write(rn, 0, label, get_lbl_fmt(BG_SUB, CLR_POS, bold=True))
            if is_pct:
                ws.write(rn, 1, val, get_pct_fmt(BG_SUB, val_color))
            else:
                ws.write(rn, 1, val, get_num_fmt(BG_SUB, val_color, bold=True))
            rn += 1

    buf.seek(0)
    return buf.read()

def build_pdf():
    from fpdf import FPDF
    pdf = FPDF(orientation="L", unit="mm", format="A3")
    pdf.add_page()
    pdf.set_margins(10, 10, 10)
    pdf.set_auto_page_break(True, margin=15)

    HDR = (254, 251, 223)   # Encabezados y subtítulos de sección (único fondo amarillo) #FEFBDF
    SUB = (255, 255, 255)   # Filas de subtotal/total -> blanco
    REG = (255, 255, 255)   # Filas regulares -> blanco
    TEXT = (51, 51, 51)     # texto neutro para encabezados

    # Formato condicional: valores numéricos negativos -> rojo, positivos -> negro.
    # Las etiquetas (texto, no numérico) siempre en negro.
    TXT_POS = (0, 0, 0)
    TXT_NEG = (220, 38, 38)

    col_w   = max(18, int(360 / (N + 2)))
    label_w = 50
    hdrs    = [f"Año {i + 1}" for i in range(len(SCOLS))] + ["SUBTOTAL"]

    ts = datetime.now().strftime("%d/%m/%Y  %H:%M")
    pdf.set_font("Helvetica", "I", 8); pdf.set_text_color(136, 136, 136)
    pdf.cell(0, 5, f"Generado: {ts}", ln=True, align="C")
    pdf.set_font("Helvetica", "B", 16); pdf.set_text_color(*TEXT)
    pdf.cell(0, 10, _pdf_safe(f"Modelo dinámico de valoración y análisis financiero - {selected.upper()}"), ln=True, align="C")
    pdf.set_text_color(*TEXT)
    pdf.ln(3)

    def draw_header():
        pdf.set_fill_color(*HDR); pdf.set_text_color(*TEXT); pdf.set_font("Helvetica", "B", 7)
        pdf.cell(label_w, 7, "Concepto", border=1, align="C", fill=True)
        for h in hdrs:
            pdf.cell(col_w, 7, h, border=1, align="C", fill=True)
        pdf.ln()

    def draw_sec_title(t):
        pdf.set_fill_color(*HDR); pdf.set_text_color(*TEXT); pdf.set_font("Helvetica", "B", 8)
        pdf.cell(label_w + col_w * len(hdrs), 6, f"  {t}", border=1, fill=True, ln=True)

    def fc(v):
        return "-" if v == 0 else (f"({abs(v):,.0f})" if v < 0 else f"{v:,.0f}")

    def draw_row(label, vals):
        row_sum = sum(vals)
        pdf.set_fill_color(*REG)
        pdf.set_text_color(*TXT_POS)
        pdf.set_font("Helvetica", "", 7)
        pdf.cell(label_w, 6, _pdf_safe(f"  {label}"), border=1, fill=True)
        for v in vals:
            pdf.set_text_color(*(TXT_NEG if v < 0 else TXT_POS))
            pdf.cell(col_w, 6, fc(v), border=1, align="R", fill=True)
        pdf.set_fill_color(*SUB)
        pdf.set_text_color(*(TXT_NEG if row_sum < 0 else TXT_POS))
        pdf.cell(col_w, 6, fc(row_sum), border=1, align="R", fill=True)
        pdf.ln()

    def draw_fcf(label, vals, sub, fill=SUB):
        pdf.set_fill_color(*fill)
        pdf.set_text_color(*TXT_POS)
        pdf.set_font("Helvetica", "B", 7)
        pdf.cell(label_w, 6, _pdf_safe(f"  {label}"), border=1, fill=True)
        for v in vals:
            pdf.set_text_color(*(TXT_NEG if v < 0 else TXT_POS))
            pdf.cell(col_w, 6, fc(v), border=1, align="R", fill=True)
        pdf.set_text_color(*(TXT_NEG if sub < 0 else TXT_POS))
        pdf.cell(col_w, 6, fc(sub), border=1, align="R", fill=True)
        pdf.ln()

    draw_header()
    draw_sec_title("INFLOWS")
    for lbl, vals in inflows:  draw_row(lbl, vals)
    draw_fcf("TOTAL INFLOWS", inflows_yr, sum(inflows_yr))
    pdf.ln(2)
    draw_sec_title("OUTFLOWS")
    for lbl, vals in outflows: draw_row(lbl, vals)
    draw_fcf("TOTAL OUTFLOWS", outflows_yr, sum(outflows_yr))
    pdf.ln(2)
    draw_fcf("DEPRECIACIÓN (Lineal)", depreciation_yr, depreciation_total)
    pdf.ln(4)
    draw_fcf("FCF PROJECT", fcf_no_fin, npv_no)
    pdf.ln(4)
    draw_sec_title("FCF FROM FINANCING")
    for lbl, vals in financing: draw_row(lbl, vals)
    draw_fcf("TOTAL FINANCING", financing_yr, sum(financing_yr))
    pdf.ln(2)
    draw_sec_title("FREE CASH FLOW")
    draw_fcf("FCF (Sin Financiamiento)", fcf_no_fin,   npv_no)
    draw_fcf("FCF (Con Financiamiento)", fcf_with_fin, npv_fin)
    pdf.ln(5)

    pdf.set_fill_color(*HDR); pdf.set_text_color(*TEXT); pdf.set_font("Helvetica", "B", 8)
    pdf.cell(130, 7, "  INVESTMENT RETURNS", border=1, fill=True, ln=True)
    for lbl, val, raw in [
        ("Tasa de Descuento",      fmt_pct(discount_rate * 100, none_str="-"), discount_rate),
        ("IRR Sin Financiamiento", fmt_pct(irr_no  * 100 if irr_no  is not None else None, none_str="-"), irr_no  or 0),
        ("IRR Con Financiamiento", fmt_pct(irr_fin * 100 if irr_fin is not None else None, none_str="-"), irr_fin or 0),
        ("NPV Sin Financiamiento", fmt_usd(van_no).replace("—", "-"),  van_no  or 0),
        ("NPV Con Financiamiento", fmt_usd(van_fin).replace("—", "-"), van_fin or 0),
        ("EBITDA",                 fmt_usd(ebitda_total).replace("—", "-"),       ebitda_total),
        ("Depreciación",           fmt_usd(depreciation_total).replace("—", "-"), depreciation_total),
        ("EBIT",                   fmt_usd(ebit_total).replace("—", "-"),         ebit_total),
    ]:
        pdf.set_fill_color(*SUB); pdf.set_font("Helvetica", "B", 7)
        pdf.set_text_color(*TXT_POS)
        pdf.cell(80, 6, f"  {lbl}", border=1, fill=True)
        pdf.set_text_color(*(TXT_NEG if raw < 0 else TXT_POS))
        pdf.cell(50, 6, val, border=1, align="R", fill=True)
        pdf.ln()

    pdf_bytes = pdf.output(dest="S")
    if isinstance(pdf_bytes, str):
        pdf_bytes = pdf_bytes.encode("latin-1")
    else:
        pdf_bytes = bytes(pdf_bytes)
    buf = BytesIO(pdf_bytes)
    buf.seek(0)
    return buf.read()


if fmt_choice == "Excel (.xlsx)":
    st.download_button(
        "⬇️ Descargar Excel",
        build_excel(),
        file_name=f"DCF_{selected}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        key=f"dl_excel_{selected}",
    )
else:
    try:
        pdf_data = build_pdf()
    except Exception as e:
        st.error(f"No se pudo generar el PDF: {e}")
    else:
        st.download_button(
            "⬇️ Descargar PDF",
            pdf_data,
            file_name=f"DCF_{selected}.pdf",
            mime="application/pdf",
            type="primary",
            key=f"dl_pdf_{selected}",
        )
